"""Ingest pipeline for student knowledge documents."""

from __future__ import annotations

import csv
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from pypdf import PdfReader
try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional until dependency is installed
    load_workbook = None

try:
    import xlrd
except ImportError:  # pragma: no cover - optional until dependency is installed
    xlrd = None

from app.core.paths import resolve_project_root

PROJECT_ROOT = resolve_project_root(Path(__file__))
DEFAULT_RAW_DIR = PROJECT_ROOT / "knowledge_raw"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "knowledge_processed"
DEFAULT_DOCUMENTS_DIR = DEFAULT_OUTPUT_DIR / "documents"
DEFAULT_CHUNKS_DIR = DEFAULT_OUTPUT_DIR / "chunks"
DEFAULT_TABLES_DIR = DEFAULT_OUTPUT_DIR / "tables"
DEFAULT_MANIFEST_PATH = DEFAULT_OUTPUT_DIR / "knowledge_manifest.csv"
SUPPORTED_DOCUMENT_EXTENSIONS = {".pdf", ".csv", ".xlsx", ".xls"}

MAX_SUMMARY_LENGTH = 240
TARGET_CHUNK_LENGTH = 1100
MAX_CHUNK_LENGTH = 1500

MANIFEST_FIELDS = [
    "file_name",
    "relative_path",
    "document_id",
    "title",
    "category",
    "page_count",
    "status",
    "source_office",
    "issued_date",
    "effective_date",
    "version",
    "source_url",
    "notes",
]


@dataclass(frozen=True)
class ExtractedPage:
    """A cleaned page extracted from a PDF."""

    page_number: int
    text: str
    paragraphs: list[str]


@dataclass(frozen=True)
class IngestedDocument:
    """The processed representation of a single PDF."""

    document_payload: dict[str, object]
    chunk_payload: dict[str, object]
    table_payload: dict[str, object] | None = None


@dataclass(frozen=True)
class IngestResult:
    """Summary returned after ingesting a knowledge directory."""

    raw_dir: Path
    output_dir: Path
    document_count: int
    chunk_count: int
    manifest_path: Path
    documents_dir: Path
    chunks_dir: Path
    tables_dir: Path


def slugify(value: str) -> str:
    """Convert a title or file name into an ASCII-safe identifier."""
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    collapsed = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_value.lower()).strip("-")
    return collapsed or "document"


def normalize_whitespace(text: str) -> str:
    """Collapse repeated whitespace while keeping paragraph breaks stable."""
    text = text.replace("\u00a0", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def infer_title_from_filename(file_name: str) -> str:
    """Build a readable title from a file name."""
    stem = Path(file_name).stem
    stem = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", stem)
    stem = re.sub(r"[_-]+", " ", stem)
    stem = re.sub(r"\s{2,}", " ", stem)
    return stem.strip()


def infer_category_name(path: Path, raw_dir: Path) -> str:
    """Infer a category label from the document parent folder."""
    try:
        relative_parent = path.parent.relative_to(raw_dir)
        parts = [part for part in relative_parent.parts if part not in {".", ""}]
    except ValueError:
        parts = [path.parent.name]

    raw_category = " / ".join(parts) if parts else "khac"
    return raw_category.replace("_", " ").replace("-", " ").strip().title()


def summarize_text(text: str, *, limit: int = MAX_SUMMARY_LENGTH) -> str:
    """Create a short preview summary from document content."""
    summary = re.sub(r"\s+", " ", text).strip()
    if len(summary) <= limit:
        return summary
    clipped = summary[:limit].rsplit(" ", 1)[0].strip()
    return f"{clipped}..."


def is_heading(paragraph: str) -> bool:
    """Heuristic for detecting section headings inside extracted text."""
    cleaned = paragraph.strip()
    if not cleaned or len(cleaned) > 160:
        return False

    normalized = cleaned.casefold()
    if cleaned.endswith(":"):
        return True
    if re.match(r"^(chuong|ph[aà]n|muc|dieu|section|article)\b", normalized):
        return True

    letters = [char for char in cleaned if char.isalpha()]
    if not letters:
        return False

    uppercase_ratio = sum(1 for char in letters if char.isupper()) / len(letters)
    return uppercase_ratio >= 0.75 and len(cleaned.split()) <= 18


def clean_page_text(text: str) -> str:
    """Clean a single extracted PDF page."""
    normalized = normalize_whitespace(text)
    lines = [line.strip() for line in normalized.split("\n")]

    paragraphs: list[str] = []
    current_parts: list[str] = []
    for line in lines:
        line = re.sub(r"^\s*STSV\s*\*\s*\d+\s*", "", line, flags=re.IGNORECASE)
        line = re.sub(r"^\s*\d+\s*\*\s*STSV\s*", "", line, flags=re.IGNORECASE)
        line = re.sub(r"\s{2,}", " ", line).strip()

        if not line:
            if current_parts:
                paragraphs.append(" ".join(current_parts).strip())
                current_parts = []
            continue

        if re.fullmatch(r"\d{1,3}", line):
            continue

        current_parts.append(line)

    if current_parts:
        paragraphs.append(" ".join(current_parts).strip())

    cleaned_paragraphs = [paragraph for paragraph in paragraphs if not is_noise_paragraph(paragraph)]
    return "\n\n".join(cleaned_paragraphs)


def is_noise_paragraph(paragraph: str) -> bool:
    """Drop obvious PDF header/footer noise."""
    cleaned = paragraph.strip()
    if not cleaned:
        return True
    if re.fullmatch(r"(stsv|\d+)(\s*\*\s*(stsv|\d+))*", cleaned, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"[\W\d_]+", cleaned):
        return True

    letter_count = sum(1 for char in cleaned if char.isalpha())
    if letter_count < 4 and len(cleaned) < 24:
        return True

    return False


def is_table_of_contents_chunk(text: str) -> bool:
    """Detect table-of-contents style chunks that should not dominate retrieval."""
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii").lower()
    dot_leader_count = len(re.findall(r"\.{4,}", text))
    short_line_count = len(re.findall(r"\b\d{1,3}\b", text))
    return "muc luc" in normalized or (dot_leader_count >= 2 and short_line_count >= 2)


def extract_pdf_pages(path: Path) -> list[ExtractedPage]:
    """Extract and clean all pages from a PDF file."""
    reader = PdfReader(str(path))
    pages: list[ExtractedPage] = []

    for index, page in enumerate(reader.pages, start=1):
        raw_text = page.extract_text() or ""
        cleaned_text = clean_page_text(raw_text)
        paragraphs = [paragraph.strip() for paragraph in cleaned_text.split("\n\n") if paragraph.strip()]
        pages.append(ExtractedPage(page_number=index, text=cleaned_text, paragraphs=paragraphs))

    return pages


def extract_layout_lines(path: Path) -> list[tuple[int, str]]:
    """Extract PDF text line-by-line using layout mode for table parsing."""
    reader = PdfReader(str(path))
    extracted_lines: list[tuple[int, str]] = []

    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text(extraction_mode="layout") or ""
        for line in text.splitlines():
            normalized_line = re.sub(r"\s+", " ", line).strip()
            if normalized_line:
                extracted_lines.append((index, normalized_line))

    return extracted_lines


def build_chunks(pages: list[ExtractedPage]) -> list[dict[str, object]]:
    """Chunk extracted pages into retrieval-friendly sections."""
    paragraphs: list[tuple[int, str]] = []
    for page in pages:
        for paragraph in page.paragraphs:
            paragraphs.append((page.page_number, paragraph))

    chunks: list[dict[str, object]] = []
    buffer: list[str] = []
    page_from: int | None = None
    page_to: int | None = None
    section_title: str | None = None

    def flush() -> None:
        nonlocal buffer, page_from, page_to, section_title
        if not buffer or page_from is None or page_to is None:
            return

        content = "\n\n".join(buffer).strip()
        if not content or len(content) < 40 or is_table_of_contents_chunk(content):
            buffer = []
            page_from = None
            page_to = None
            section_title = None
            return

        chunks.append(
            {
                "section_title": section_title,
                "summary": summarize_text(content),
                "content": content,
                "page_from": page_from,
                "page_to": page_to,
            }
        )

        buffer = []
        page_from = None
        page_to = None
        section_title = None

    for paragraph_page, paragraph in paragraphs:
        paragraph_length = len(paragraph)
        current_length = sum(len(item) for item in buffer) + (2 * len(buffer))
        starts_new_section = is_heading(paragraph) and buffer
        would_overflow = current_length + paragraph_length > MAX_CHUNK_LENGTH
        reached_target = buffer and current_length >= TARGET_CHUNK_LENGTH

        if starts_new_section or would_overflow or reached_target:
            flush()

        if page_from is None:
            page_from = paragraph_page
        page_to = paragraph_page

        if section_title is None and is_heading(paragraph):
            section_title = paragraph

        buffer.append(paragraph)

    flush()
    return chunks


def _normalize_amount_text(amount_text: str) -> str:
    """Normalize OCR variants inside an extracted amount string."""
    normalized = amount_text.strip()
    normalized = normalized.replace("d6ng", "dong")
    normalized = normalized.replace("ddng", "dong")
    normalized = normalized.replace("đ6ng", "dong")
    normalized = normalized.replace("đồng", "dong")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def _normalize_ocr_phrase(text: str) -> str:
    """Normalize common OCR artifacts in tuition table titles and labels."""
    normalized = text.strip()
    replacements = {
        "chatA luong": "chat luong",
        "chat luong": "chat luong",
        "vira": "vua",
        "ngw": "ngu",
        "mire": "muc",
        "tii": "tu",
        "ttr": "tu",
        "tii'": "tu",
        "trdâ€™ve": "tro ve",
        "trd've": "tro ve",
        "quoc phong": "quoc phong",
        "giao due": "giao duc",
        "chinh quy": "chinh quy",
    }
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def _amount_to_value(amount_text: str) -> int | None:
    """Convert an amount like 1.473.287 dong to an integer value."""
    match = re.search(r"(\d[\d\.]*)", amount_text)
    if not match:
        return None

    digits = match.group(1).replace(".", "")
    return int(digits) if digits.isdigit() else None


def _infer_tuition_track_tags(title: str, label: str, amount_text: str) -> tuple[list[str], list[str]]:
    """Infer structured tags for tuition rows from OCR-extracted text."""
    combined = slugify(" ".join([title, label, amount_text]))
    track_tags: list[str] = []
    basis_tags: list[str] = []

    def add_tag(target: list[str], value: str) -> None:
        if value not in target:
            target.append(value)

    if "cao-hoc" in combined or "thac-si" in combined:
        add_tag(track_tags, "cao_hoc")
    if "chat-luong-cao" in combined or "clc" in combined or "tien-tien" in combined:
        add_tag(track_tags, "chat_luong_cao")
    if "tien-tien" in combined:
        add_tag(track_tags, "tien_tien")
    if "lien-thong" in combined:
        add_tag(track_tags, "lien_thong")
    if "bang-2" in combined:
        add_tag(track_tags, "bang_2")
    if "vua-lam-vua-hoc" in combined:
        add_tag(track_tags, "vua_lam_vua_hoc")
    if "giao-duc-the-chat" in combined or "quoc-phong" in combined:
        add_tag(track_tags, "the_chat_quoc_phong")
    if not track_tags and "chinh-quy" in combined:
        add_tag(track_tags, "chinh_quy")
    if not track_tags:
        add_tag(track_tags, "chinh_quy")

    if "/-1-tin-chi" in combined or "/-hoc" in combined or "theo-tin-chi" in combined:
        add_tag(basis_tags, "per_credit")
    if "dao-tao-dai-hoc-he-chinh-quy" in combined:
        add_tag(basis_tags, "per_credit")
    if "chat-luong-cao" in combined or "lien-thong" in combined or "bang-2" in combined:
        add_tag(basis_tags, "per_credit")
    if "cao-hoc" in combined and ("1-1" in combined or "1-2" in combined or "1-3" in combined):
        add_tag(basis_tags, "per_credit")
    if "cao-hoc" in combined and not basis_tags:
        add_tag(basis_tags, "full_program")
    if not basis_tags and _amount_to_value(amount_text) and (_amount_to_value(amount_text) or 0) >= 10_000_000:
        add_tag(basis_tags, "full_program")

    return track_tags, basis_tags


def _build_tuition_search_text(title: str, label: str, amount_text: str, track_tags: list[str], basis_tags: list[str]) -> str:
    """Create a normalized search text for structured tuition rows."""
    alias_tokens: list[str] = ["hoc phi"]
    alias_map = {
        "chinh_quy": ["chinh quy", "dai hoc"],
        "chat_luong_cao": ["chat luong cao", "clc"],
        "tien_tien": ["tien tien"],
        "lien_thong": ["lien thong"],
        "bang_2": ["bang 2"],
        "vua_lam_vua_hoc": ["vua lam vua hoc"],
        "the_chat_quoc_phong": ["giao duc the chat", "quoc phong"],
        "cao_hoc": ["cao hoc", "thac si"],
    }
    basis_map = {
        "per_credit": ["tin chi", "1 tin chi"],
        "full_program": ["toan khoa", "tron khoa"],
    }

    for tag in track_tags:
        alias_tokens.extend(alias_map.get(tag, []))
    for tag in basis_tags:
        alias_tokens.extend(basis_map.get(tag, []))

    alias_text = " ".join(alias_tokens)
    return _normalize_ocr_phrase(f"{title} {label} {amount_text} {alias_text}")


def _is_table_row_candidate(line: str) -> bool:
    """Detect whether a layout line looks like a table row."""
    return bool(re.match(r"^(?:\d+(?:\.\d+)?|i)\s+", line, flags=re.IGNORECASE))


def _is_table_header_line(line: str) -> bool:
    """Drop obvious column headers and decorative OCR fragments."""
    normalized = slugify(line)
    return (
        "tt" == normalized
        or "khoa-khoi-nganh" in normalized
        or "muc-hoc-phi" in normalized
        or normalized in {"1-tin-chi", "hoc-ky"}
    )


def _is_table_section_heading(line: str) -> bool:
    """Detect section headings like I., II., IX."""
    match = re.match(r"^(?P<roman>(?:[IVXLCDM]{1,6}|IL))\.?\s*(?P<title>.+)$", line, flags=re.IGNORECASE)
    if not match:
        return False

    title = match.group("title").strip()
    if len(title) < 8:
        return False

    normalized_title = slugify(title)
    return any(
        keyword in normalized_title
        for keyword in ("muc-thu", "hoc-phi", "dao-tao", "chuong-trinh", "nganh")
    )


def _parse_table_row(line: str) -> tuple[str, str, str] | None:
    """Parse a table row into row id, label, and amount."""
    cleaned = line.strip()
    match = re.match(
        r"^(?P<row_id>\d+(?:\.\d+)?|i)\s+(?P<body>.+?)\s+(?P<amount>\d[\d\.]*\s*(?:dong|ddng|d6ng)(?:/\s*(?:1|hoc))?)$",
        cleaned,
        flags=re.IGNORECASE,
    )
    if match:
        row_id = match.group("row_id")
        label = re.sub(r"\s+", " ", match.group("body")).strip(" -")
        amount = _normalize_amount_text(match.group("amount"))
        return row_id, label, amount

    if not re.search(r"\d[\d\.]*", cleaned):
        return None

    row_prefix = re.match(r"^(?P<row_id>\d+(?:\.\d+)?|i)\s+", cleaned, flags=re.IGNORECASE)
    if not row_prefix:
        return None

    row_id = row_prefix.group("row_id")
    body = cleaned[row_prefix.end() :].strip()
    amount_match = re.search(
        r"(?P<amount>\d[\d\.]*\s*(?:dong|ddng|d6ng)(?:/\s*(?:1|hoc))?)",
        body,
        flags=re.IGNORECASE,
    )
    if not amount_match:
        leading_amount_match = re.match(
            r"^(?P<amount>\d[\d\.]*\s*(?:dong|ddng|d6ng))(?P<label>.+)$",
            body,
            flags=re.IGNORECASE,
        )
        if not leading_amount_match:
            return None

        amount = _normalize_amount_text(leading_amount_match.group("amount"))
        label = re.sub(r"\s+", " ", leading_amount_match.group("label")).strip(" -")
        return row_id, label, amount

    if len(re.sub(r"\D", "", amount_match.group("amount"))) < 5:
        return None

    amount = _normalize_amount_text(amount_match.group("amount"))
    label = re.sub(r"\s+", " ", body[: amount_match.start()]).strip(" -")
    trailing = re.sub(r"\s+", " ", body[amount_match.end() :]).strip(" -")
    if trailing:
        label = f"{label} {trailing}".strip()

    return row_id, label, amount


def extract_structured_tables(
    path: Path,
    *,
    document_id: str,
    title: str,
    category: str,
    source_path: str,
) -> dict[str, object] | None:
    """Extract structured tables from tuition PDFs using layout-mode text."""
    if path.parent.name.casefold() != "hoc_phi":
        return None

    lines = extract_layout_lines(path)
    if not lines:
        return None

    tables: list[dict[str, object]] = []
    current_table: dict[str, object] | None = None
    current_row: dict[str, object] | None = None

    def flush_row() -> None:
        nonlocal current_row, current_table
        if current_table is None or current_row is None:
            return

        label = re.sub(r"\s+", " ", str(current_row["label"])).strip(" /-")
        amount_text = _normalize_amount_text(str(current_row["amount_text"]))
        if label and amount_text:
            table_title = _normalize_ocr_phrase(str(current_table["title"]))
            clean_label = _normalize_ocr_phrase(label)
            track_tags, basis_tags = _infer_tuition_track_tags(table_title, clean_label, amount_text)
            current_row["label"] = clean_label
            current_row["amount_text"] = amount_text
            current_row["amount_value"] = _amount_to_value(amount_text)
            current_row["track_tags"] = track_tags
            current_row["basis_tags"] = basis_tags
            current_row["search_text"] = _build_tuition_search_text(
                table_title,
                clean_label,
                amount_text,
                track_tags,
                basis_tags,
            )
            current_table["rows"].append(current_row)

        current_row = None

    def flush_table() -> None:
        nonlocal current_table
        flush_row()
        if current_table and current_table["rows"]:
            tables.append(current_table)
        current_table = None

    for page_number, line in lines:
        if _is_table_section_heading(line):
            flush_table()
            heading = _normalize_ocr_phrase(re.sub(r"\s+", " ", line).strip())
            current_table = {
                "table_id": f"{document_id}-table-{len(tables) + 1:02d}",
                "title": heading,
                "page_from": page_number,
                "page_to": page_number,
                "rows": [],
            }
            continue

        if current_table is None:
            continue

        current_table["page_to"] = page_number

        if re.match(r"^(ghi chu|phu luc)\b", line, flags=re.IGNORECASE):
            flush_table()
            continue

        if _is_table_header_line(line):
            continue

        if current_row is None and not current_table["rows"]:
            normalized_line = slugify(line)
            if normalized_line and not _is_table_row_candidate(line):
                alpha_tokens = [token for token in normalized_line.split("-") if token.isalpha()]
                if len(alpha_tokens) < 2 or sum(len(token) for token in alpha_tokens) < 6:
                    continue
                if not any(token in normalized_line for token in ("tt", "khoa", "muc-hoc-phi", "1-tin-chi")):
                    current_table["title"] = _normalize_ocr_phrase(
                        f"{current_table['title']} {line}".strip()
                    )
                continue

        parsed_row = _parse_table_row(line)
        if parsed_row:
            flush_row()
            row_id, label, amount_text = parsed_row
            current_row = {
                "row_id": row_id,
                "label": label,
                "amount_text": amount_text,
                "amount_value": _amount_to_value(amount_text),
                "page_from": page_number,
                "page_to": page_number,
            }
            continue

        if current_row is not None:
            if re.search(r"^(tin chi|hoc ky)$", line, flags=re.IGNORECASE):
                current_row["amount_text"] = f"{current_row['amount_text']} {line}".strip()
                current_row["page_to"] = page_number
                continue

            if not _is_table_header_line(line):
                current_row["label"] = f"{current_row['label']} {line}".strip()
                current_row["page_to"] = page_number

    flush_table()

    if not tables:
        return None

    return {
        "document_id": document_id,
        "title": title,
        "category": category,
        "source_path": source_path,
        "source_url": "",
        "tables": tables,
    }


def _normalize_header_row(row: list[str]) -> list[str]:
    """Return stable spreadsheet headers for a parsed row."""
    normalized_headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(row, start=1):
        header = normalize_whitespace(value)
        if not header:
            header = f"column_{index}"
        key = slugify(header).replace("-", "_") or f"column_{index}"
        seen[key] = seen.get(key, 0) + 1
        if seen[key] > 1:
            key = f"{key}_{seen[key]}"
        normalized_headers.append(key)
    return normalized_headers


def _stringify_cell(value: object) -> str:
    """Normalize a spreadsheet cell into plain text."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return normalize_whitespace(str(value))


def _read_csv_rows(path: Path) -> list[list[str]]:
    """Read a CSV file into normalized rows."""
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                reader = csv.reader(file)
                rows = [[_stringify_cell(cell) for cell in row] for row in reader]
            return [row for row in rows if any(cell for cell in row)]
        except UnicodeDecodeError:
            continue
    return []


def _read_xlsx_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    """Read an XLSX workbook into sheet rows."""
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to ingest .xlsx files")

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheets: list[tuple[str, list[list[str]]]] = []
    for worksheet in workbook.worksheets:
        rows: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            cleaned = [_stringify_cell(cell) for cell in row]
            if any(cleaned):
                rows.append(cleaned)
        if rows:
            sheets.append((worksheet.title, rows))
    return sheets


def _read_xls_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    """Read an XLS workbook into sheet rows."""
    if xlrd is None:
        raise RuntimeError("xlrd is required to ingest .xls files")

    workbook = xlrd.open_workbook(path)
    sheets: list[tuple[str, list[list[str]]]] = []
    for sheet_index in range(workbook.nsheets):
        worksheet = workbook.sheet_by_index(sheet_index)
        rows: list[list[str]] = []
        for row_index in range(worksheet.nrows):
            cleaned = [_stringify_cell(worksheet.cell_value(row_index, column_index)) for column_index in range(worksheet.ncols)]
            if any(cleaned):
                rows.append(cleaned)
        if rows:
            sheets.append((worksheet.name, rows))
    return sheets


def _rows_to_structured_table(
    *,
    document_id: str,
    title: str,
    category: str,
    source_path: str,
    sheet_name: str,
    rows: list[list[str]],
    table_index: int,
) -> dict[str, object] | None:
    """Convert spreadsheet rows into a table payload."""
    if len(rows) < 2:
        return None

    header_row = rows[0]
    data_rows = rows[1:]
    headers = _normalize_header_row(header_row)
    sheet_title = sheet_name.strip() or f"Sheet {table_index}"
    table_title = f"{title} - {sheet_title}" if sheet_title.lower() != title.lower() else title

    table_rows: list[dict[str, object]] = []
    for row_index, raw_row in enumerate(data_rows, start=1):
        padded = list(raw_row) + [""] * max(0, len(headers) - len(raw_row))
        values = padded[: len(headers)]
        row_data = {header: value for header, value in zip(headers, values, strict=False)}
        pairs = [(header, value) for header, value in row_data.items() if value]
        if not pairs:
            continue

        label = " - ".join(value for _, value in pairs[:2]).strip() or f"Row {row_index}"
        amount_text = " | ".join(f"{header}: {value}" for header, value in pairs[2:]).strip() or None
        search_text = " ".join([table_title, *[f"{header} {value}" for header, value in pairs]]).strip()

        track_tags: list[str] = []
        basis_tags: list[str] = []
        if "hoc_phi" in slugify(source_path):
            inferred_track_tags, inferred_basis_tags = _infer_tuition_track_tags(table_title, label, amount_text or "")
            track_tags = inferred_track_tags
            basis_tags = inferred_basis_tags

        table_rows.append(
            {
                "row_id": f"row-{row_index:03d}",
                "label": label,
                "amount_text": amount_text,
                "search_text": search_text,
                "track_tags": track_tags,
                "basis_tags": basis_tags,
                "row_data": row_data,
                "page_from": table_index,
                "page_to": table_index,
            }
        )

    if not table_rows:
        return None

    return {
        "table_id": f"{document_id}-table-{table_index:02d}",
        "title": table_title,
        "headers": headers,
        "page_from": table_index,
        "page_to": table_index,
        "rows": table_rows,
    }


def ingest_spreadsheet(path: Path, raw_dir: Path) -> IngestedDocument:
    """Ingest one spreadsheet file into document, chunk, and table payloads."""
    category = infer_category_name(path, raw_dir)
    title = infer_title_from_filename(path.name)
    document_id = slugify(f"{path.parent.name}-{path.stem}")
    source_path = str(path.relative_to(PROJECT_ROOT)).replace("\\", "/")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        sheet_rows = [(path.stem, _read_csv_rows(path))]
    elif suffix == ".xlsx":
        sheet_rows = _read_xlsx_sheets(path)
    elif suffix == ".xls":
        sheet_rows = _read_xls_sheets(path)
    else:
        raise ValueError(f"Unsupported spreadsheet format: {suffix}")

    pages: list[ExtractedPage] = []
    tables: list[dict[str, object]] = []
    for sheet_index, (sheet_name, rows) in enumerate(sheet_rows, start=1):
        if not rows:
            continue

        table = _rows_to_structured_table(
            document_id=document_id,
            title=title,
            category=category,
            source_path=source_path,
            sheet_name=sheet_name,
            rows=rows,
            table_index=sheet_index,
        )
        if table is not None:
            tables.append(table)

        header_names = _normalize_header_row(rows[0]) if rows else []
        paragraphs = [
            f"Bang du lieu {sheet_name}. Cac cot: {', '.join(header_names)}."
        ]
        for row in rows[1:]:
            values = row[: len(header_names)]
            pairs = [(header, value) for header, value in zip(header_names, values, strict=False) if value]
            if pairs:
                paragraphs.append(". ".join(f"{header}: {value}" for header, value in pairs))

        full_sheet_text = "\n\n".join(paragraphs).strip()
        pages.append(
            ExtractedPage(
                page_number=sheet_index,
                text=full_sheet_text,
                paragraphs=paragraphs,
            )
        )

    full_text = "\n\n".join(page.text for page in pages if page.text).strip()
    summary = summarize_text(full_text)
    chunks = build_chunks(pages)
    chunk_payloads: list[dict[str, object]] = []
    for index, chunk in enumerate(chunks, start=1):
        chunk_payloads.append(
            {
                "chunk_id": f"{document_id}-chunk-{index:03d}",
                "section_title": chunk["section_title"],
                "summary": chunk["summary"],
                "content": chunk["content"],
                "page_from": chunk["page_from"],
                "page_to": chunk["page_to"],
                "keywords": [category.casefold(), title.casefold(), "bang du lieu"],
            }
        )

    document_payload = {
        "id": document_id,
        "title": title,
        "category": category,
        "summary": summary,
        "content": full_text,
        "source_url": "",
        "source_path": source_path,
        "source_file": path.name,
        "page_count": len(pages),
        "status": "needs_review",
        "source_office": "",
        "issued_date": "",
        "effective_date": "",
        "version": "",
        "keywords": [category.casefold(), title.casefold(), "bang du lieu"],
    }

    chunk_payload = {
        "document_id": document_id,
        "title": title,
        "category": category,
        "source_url": "",
        "source_path": source_path,
        "page_count": len(pages),
        "status": "needs_review",
        "chunks": chunk_payloads,
    }

    table_payload = {
        "document_id": document_id,
        "title": title,
        "category": category,
        "source_path": source_path,
        "source_url": "",
        "tables": tables,
    }

    return IngestedDocument(
        document_payload=document_payload,
        chunk_payload=chunk_payload,
        table_payload=table_payload if tables else None,
    )


def ingest_document(path: Path, raw_dir: Path) -> IngestedDocument:
    """Dispatch ingestion based on the file extension."""
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return ingest_pdf(path, raw_dir)
    if suffix in {".csv", ".xlsx", ".xls"}:
        return ingest_spreadsheet(path, raw_dir)
    raise ValueError(f"Unsupported knowledge document format: {suffix}")


def ingest_pdf(path: Path, raw_dir: Path) -> IngestedDocument:
    """Ingest one PDF file into document and chunk payloads."""
    category = infer_category_name(path, raw_dir)
    title = infer_title_from_filename(path.name)
    document_id = slugify(f"{path.parent.name}-{path.stem}")
    source_path = str(path.relative_to(PROJECT_ROOT)).replace("\\", "/")

    pages = extract_pdf_pages(path)
    full_text = "\n\n".join(page.text for page in pages if page.text).strip()
    summary = summarize_text(full_text)

    chunks = build_chunks(pages)
    chunk_payloads: list[dict[str, object]] = []
    for index, chunk in enumerate(chunks, start=1):
        chunk_payload = {
            "chunk_id": f"{document_id}-chunk-{index:03d}",
            "section_title": chunk["section_title"],
            "summary": chunk["summary"],
            "content": chunk["content"],
            "page_from": chunk["page_from"],
            "page_to": chunk["page_to"],
            "keywords": [category.casefold(), title.casefold()],
        }
        chunk_payloads.append(chunk_payload)

    document_payload = {
        "id": document_id,
        "title": title,
        "category": category,
        "summary": summary,
        "content": full_text,
        "source_url": "",
        "source_path": source_path,
        "source_file": path.name,
        "page_count": len(pages),
        "status": "needs_review",
        "source_office": "",
        "issued_date": "",
        "effective_date": "",
        "version": "",
        "keywords": [category.casefold(), title.casefold()],
    }

    chunk_payload = {
        "document_id": document_id,
        "title": title,
        "category": category,
        "source_url": "",
        "source_path": source_path,
        "page_count": len(pages),
        "status": "needs_review",
        "chunks": chunk_payloads,
    }

    table_payload = extract_structured_tables(
        path,
        document_id=document_id,
        title=title,
        category=category,
        source_path=source_path,
    )

    return IngestedDocument(
        document_payload=document_payload,
        chunk_payload=chunk_payload,
        table_payload=table_payload,
    )


def write_manifest(path: Path, documents: list[dict[str, object]]) -> None:
    """Write a manifest CSV summarizing all ingested PDFs."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=MANIFEST_FIELDS)
        writer.writeheader()
        for document in documents:
            writer.writerow(
                {
                    "file_name": document["source_file"],
                    "relative_path": document["source_path"],
                    "document_id": document["id"],
                    "title": document["title"],
                    "category": document["category"],
                    "page_count": document["page_count"],
                    "status": document["status"],
                    "source_office": document["source_office"],
                    "issued_date": document["issued_date"],
                    "effective_date": document["effective_date"],
                    "version": document["version"],
                    "source_url": document["source_url"],
                    "notes": "",
                }
            )


def ingest_knowledge_base(
    raw_dir: Path = DEFAULT_RAW_DIR,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> IngestResult:
    """Ingest every supported document inside the raw knowledge directory."""
    raw_dir = raw_dir.resolve()
    output_dir = output_dir.resolve()
    documents_dir = output_dir / "documents"
    chunks_dir = output_dir / "chunks"
    tables_dir = output_dir / "tables"

    documents_dir.mkdir(parents=True, exist_ok=True)
    chunks_dir.mkdir(parents=True, exist_ok=True)
    tables_dir.mkdir(parents=True, exist_ok=True)

    for artifact_dir in (documents_dir, chunks_dir, tables_dir):
        for artifact_path in artifact_dir.glob("*.json"):
            artifact_path.unlink()

    supported_files = sorted(
        path for path in raw_dir.rglob("*") if path.is_file() and path.suffix.lower() in SUPPORTED_DOCUMENT_EXTENSIONS
    )
    ingested_documents: list[dict[str, object]] = []
    total_chunks = 0

    for source_path in supported_files:
        ingested = ingest_document(source_path, raw_dir)
        document = ingested.document_payload
        chunks = ingested.chunk_payload

        document_path = documents_dir / f"{document['id']}.json"
        chunk_path = chunks_dir / f"{document['id']}.json"
        document_path.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
        chunk_path.write_text(json.dumps(chunks, ensure_ascii=False, indent=2), encoding="utf-8")

        if ingested.table_payload:
            table_path = tables_dir / f"{document['id']}.json"
            table_path.write_text(
                json.dumps(ingested.table_payload, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

        ingested_documents.append(document)
        total_chunks += len(chunks["chunks"])

    manifest_path = output_dir / DEFAULT_MANIFEST_PATH.name
    write_manifest(manifest_path, ingested_documents)

    return IngestResult(
        raw_dir=raw_dir,
        output_dir=output_dir,
        document_count=len(ingested_documents),
        chunk_count=total_chunks,
        manifest_path=manifest_path,
        documents_dir=documents_dir,
        chunks_dir=chunks_dir,
        tables_dir=tables_dir,
    )
