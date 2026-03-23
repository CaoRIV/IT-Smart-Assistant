"""Workbook type detection and sheet schema metadata for Excel knowledge files."""

from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


COURSE_CATALOG_REQUIRED_HEADERS = {"ten hoc phan", "ma hoc phan", "so tc"}


def _normalize(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_value.lower().replace("_", " ").replace("-", " ").strip()


def _sheet_role(sheet_name: str) -> str:
    normalized = _normalize(sheet_name)
    if normalized == "du lieu":
        return "course_catalog_flat"
    if normalized == "tong quan":
        return "course_catalog_summary"
    if normalized in {"cu nhan", "ky su"}:
        return "course_catalog_semester_layout"
    if normalized == "ghi chu":
        return "course_catalog_glossary"
    return "generic_sheet"


def _load_sheet_headers(path: Path) -> dict[str, list[str]]:
    if load_workbook is None or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {}

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    headers_by_sheet: dict[str, list[str]] = {}
    for worksheet in workbook.worksheets:
        first_non_empty: list[str] | None = None
        for row in worksheet.iter_rows(values_only=True):
            cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(cleaned):
                first_non_empty = cleaned
                break
        if first_non_empty:
            headers_by_sheet[worksheet.title] = first_non_empty
    return headers_by_sheet


def detect_workbook_schema(path: Path, *, category: str | None = None) -> dict[str, Any] | None:
    """Detect workbook type and sheet schema config for supported spreadsheet patterns."""
    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xls", ".csv"}:
        return None

    normalized_category = _normalize(category or "")
    if suffix == ".csv":
        return {
            "workbook_type": "generic_tabular",
            "sheet_schema_config": {
                "canonical_sheet": path.stem,
                "sheet_roles": {path.stem: "generic_sheet"},
            },
        }

    headers_by_sheet = _load_sheet_headers(path)
    normalized_headers = {
        sheet_name: {_normalize(header) for header in headers if header}
        for sheet_name, headers in headers_by_sheet.items()
    }

    sheet_names = {_normalize(sheet_name) for sheet_name in headers_by_sheet}
    has_course_catalog_layout = {"tong quan", "du lieu", "ghi chu"}.issubset(sheet_names)
    canonical_sheet = None
    for sheet_name, headers in normalized_headers.items():
        if COURSE_CATALOG_REQUIRED_HEADERS.issubset(headers):
            canonical_sheet = sheet_name
            break

    if canonical_sheet and (
        has_course_catalog_layout
        or "chuong trinh dao tao" in normalized_category
        or "danh sach mon hoc" in _normalize(path.stem)
    ):
        return {
            "workbook_type": "course_catalog",
            "sheet_schema_config": {
                "canonical_sheet": canonical_sheet,
                "sheet_roles": {
                    sheet_name: _sheet_role(sheet_name) for sheet_name in headers_by_sheet
                },
            },
        }

    return {
        "workbook_type": "generic_tabular",
        "sheet_schema_config": {
            "canonical_sheet": next(iter(headers_by_sheet.keys()), path.stem),
            "sheet_roles": {
                sheet_name: "generic_sheet" for sheet_name in headers_by_sheet
            },
        },
    }
