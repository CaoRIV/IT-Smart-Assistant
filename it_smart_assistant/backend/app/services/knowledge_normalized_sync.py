"""Sync normalized knowledge layer tables from processed artifacts and admin content."""

from __future__ import annotations

import csv
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.agents.tools.form_generator import _find_admin_form
from app.agents.tools.procedure_workflow import PROCEDURE_CATALOG
from app.core.paths import resolve_project_root
from app.db.models.knowledge import KnowledgeSource, KnowledgeSourceVersion
from app.knowledge.admin_store import FAQS_DIR, FORMS_DIR, list_faqs, list_forms, load_document_metadata_map
from app.knowledge.ingest import DEFAULT_OUTPUT_DIR, DEFAULT_RAW_DIR, slugify
from app.knowledge.workbook_schema import detect_workbook_schema
from app.repositories import knowledge as knowledge_repo

PROJECT_ROOT = resolve_project_root(Path(__file__))
CHUNKS_DIR = DEFAULT_OUTPUT_DIR / "chunks"
TABLES_DIR = DEFAULT_OUTPUT_DIR / "tables"
MANIFEST_PATH = DEFAULT_OUTPUT_DIR / "knowledge_manifest.csv"


async def _normalized_tables_ready(db: AsyncSession) -> bool:
    result = await db.execute(
        text(
            """
            SELECT
                to_regclass('public.knowledge_chunks') IS NOT NULL
                AND to_regclass('public.knowledge_tables') IS NOT NULL
                AND to_regclass('public.knowledge_table_rows') IS NOT NULL
                AND to_regclass('public.knowledge_faqs') IS NOT NULL
                AND to_regclass('public.knowledge_interactions') IS NOT NULL
                AND to_regclass('public.knowledge_procedures') IS NOT NULL
                AND to_regclass('public.knowledge_forms') IS NOT NULL
                AND to_regclass('public.knowledge_course_catalogs') IS NOT NULL
                AND to_regclass('public.knowledge_courses') IS NOT NULL
            """
        )
    )
    return bool(result.scalar())


def _manifest_by_document_id() -> dict[str, dict[str, str]]:
    if not MANIFEST_PATH.exists():
        return {}

    with MANIFEST_PATH.open("r", encoding="utf-8", newline="") as file:
        reader = csv.DictReader(file)
        return {row["document_id"]: row for row in reader if row.get("document_id")}


def _source_id_from_relative_path(relative_path: str) -> str:
    return f"src-{slugify(relative_path)}"


async def _active_source_versions_by_source_id(db: AsyncSession) -> dict[str, object]:
    result = await db.execute(
        select(KnowledgeSource.source_id, KnowledgeSourceVersion.id)
        .join(KnowledgeSourceVersion, KnowledgeSourceVersion.source_pk_id == KnowledgeSource.id)
        .where(KnowledgeSourceVersion.is_active.is_(True))
    )
    return {source_id: version_id for source_id, version_id in result.all()}


def _resolve_document_context(
    document_id: str,
    source_path: str | None,
    *,
    manifest_lookup: dict[str, dict[str, str]],
    metadata_lookup: dict[str, dict[str, str]],
) -> tuple[dict[str, str], dict[str, str], str | None, str]:
    manifest_row = manifest_lookup.get(document_id, {})
    relative_path = manifest_row.get("relative_path") or source_path
    metadata_row = metadata_lookup.get(relative_path or "", {})
    status = metadata_row.get("status") or manifest_row.get("status") or "draft"
    return manifest_row, metadata_row, relative_path, status


def _workbook_type_for(relative_path: str | None, metadata_row: dict[str, str]) -> str | None:
    workbook_type = metadata_row.get("workbook_type")
    if workbook_type:
        return str(workbook_type)
    if not relative_path:
        return None
    path = PROJECT_ROOT / relative_path
    if not path.exists():
        return None
    detected = detect_workbook_schema(path, category=metadata_row.get("category"))
    return detected.get("workbook_type") if detected else None


def _safe_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _normalize_search_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


COURSE_ALIAS_REPLACEMENTS = {
    "ai so tuyen tinh": "dai so tuyen tinh",
    "tin hoc ai cuong": "tin hoc dai cuong",
    "chuyen e": "chuyen de",
    "mac le nin": "mac lenin",
}


def _normalize_course_name(value: str) -> tuple[str, list[str]]:
    normalized = _normalize_search_text(value)
    aliases: list[str] = []
    for source, target in COURSE_ALIAS_REPLACEMENTS.items():
        if source in normalized:
            aliases.append(target)
            normalized = normalized.replace(source, target)
    aliases = sorted(set(alias for alias in aliases if alias and alias != normalized))
    return normalized, aliases


def _parse_int(value: object) -> int | None:
    text_value = _safe_cell(value)
    if not text_value:
        return None
    digits = re.sub(r"[^\d-]", "", text_value)
    if not digits or digits == "-":
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def _parse_program_track(value: str) -> tuple[str, str | None]:
    normalized = _normalize_search_text(value)
    if "ky su" in normalized:
        return "ky_su", "ky_su"
    if "cu nhan" in normalized:
        return "cu_nhan", "cu_nhan"
    if "tich hop" in normalized:
        return "tich_hop", None
    return "khac", None


def _parse_teaching_language(value: str) -> str | None:
    normalized = _normalize_search_text(value)
    if not normalized:
        return None
    if "anh" in normalized:
        return "anh"
    if "viet" in normalized:
        return "viet"
    if "song ngu" in normalized:
        return "song_ngu"
    return normalized


def _parse_semester_number(value: str) -> int | None:
    match = re.search(r"(\d+)", _normalize_search_text(value))
    return int(match.group(1)) if match else None


def _looks_like_course_catalog_sheet(headers: list[str]) -> bool:
    normalized_headers = {_normalize_search_text(header) for header in headers if header}
    required = {"ten hoc phan", "ma hoc phan", "so tc"}
    return required.issubset(normalized_headers)


def _sheet_rows(worksheet) -> list[list[object]]:
    rows: list[list[object]] = []
    for row in worksheet.iter_rows(values_only=True):
        cleaned = [_safe_cell(cell) for cell in row]
        if any(cleaned):
            rows.append(cleaned)
    return rows


def _course_catalog_source_files(metadata_lookup: dict[str, dict[str, str]]) -> list[Path]:
    files: list[Path] = []
    for path in sorted(
        item
        for item in DEFAULT_RAW_DIR.rglob("*")
        if item.is_file() and item.suffix.lower() in {".xlsx", ".xlsm"}
    ):
        relative_path = str(path.relative_to(PROJECT_ROOT)).replace("\\", "/")
        if _workbook_type_for(relative_path, metadata_lookup.get(relative_path, {})) == "course_catalog":
            files.append(path)
    return files


async def _sync_chunks(
    db: AsyncSession,
    *,
    manifest_lookup: dict[str, dict[str, str]],
    metadata_lookup: dict[str, dict[str, str]],
    source_version_lookup: dict[str, object],
) -> int:
    count = 0
    if not CHUNKS_DIR.exists():
        return count

    for path in sorted(CHUNKS_DIR.glob("*.json")):
        raw = json.loads(path.read_text(encoding="utf-8"))
        document_id = raw.get("document_id")
        if not document_id:
            continue

        manifest_row, metadata_row, relative_path, status = _resolve_document_context(
            document_id,
            raw.get("source_path"),
            manifest_lookup=manifest_lookup,
            metadata_lookup=metadata_lookup,
        )
        if _workbook_type_for(relative_path, metadata_row) == "course_catalog":
            continue
        source_version_id = (
            source_version_lookup.get(_source_id_from_relative_path(relative_path))
            if relative_path
            else None
        )

        for chunk in raw.get("chunks", []):
            await knowledge_repo.create_knowledge_chunk(
                db,
                payload={
                    "chunk_id": chunk["chunk_id"],
                    "source_version_id": source_version_id,
                    "heading": chunk.get("section_title"),
                    "heading_level": None,
                    "section_title": chunk.get("section_title"),
                    "content": chunk.get("content") or "",
                    "summary": chunk.get("summary"),
                    "page_from": chunk.get("page_from"),
                    "page_to": chunk.get("page_to"),
                    "keywords": chunk.get("keywords", []),
                    "metadata_json": {
                        "document_id": document_id,
                        "title": raw.get("title"),
                        "category": metadata_row.get("category") or raw.get("category"),
                        "source_path": raw.get("source_path"),
                        "relative_path": relative_path,
                        "source_url": metadata_row.get("source_url") or raw.get("source_url"),
                        "source_office": metadata_row.get("source_office") or manifest_row.get("source_office"),
                    },
                    "status": status,
                },
            )
            count += 1

    return count


async def _sync_tables(
    db: AsyncSession,
    *,
    manifest_lookup: dict[str, dict[str, str]],
    metadata_lookup: dict[str, dict[str, str]],
    source_version_lookup: dict[str, object],
) -> tuple[int, int]:
    table_count = 0
    row_count = 0
    if not TABLES_DIR.exists():
        return table_count, row_count

    for path in sorted(TABLES_DIR.glob("*.json")):
        raw = json.loads(path.read_text(encoding="utf-8"))
        document_id = raw.get("document_id")
        if not document_id:
            continue

        manifest_row, metadata_row, relative_path, status = _resolve_document_context(
            document_id,
            raw.get("source_path"),
            manifest_lookup=manifest_lookup,
            metadata_lookup=metadata_lookup,
        )
        if _workbook_type_for(relative_path, metadata_row) == "course_catalog":
            continue
        source_version_id = (
            source_version_lookup.get(_source_id_from_relative_path(relative_path))
            if relative_path
            else None
        )

        for table in raw.get("tables", []):
            db_table = await knowledge_repo.create_knowledge_table(
                db,
                payload={
                    "table_id": table["table_id"],
                    "source_version_id": source_version_id,
                    "title": table.get("title") or raw.get("title") or "Bang du lieu",
                    "schema_type": "tuition_table" if "hoc_phi" in (relative_path or "") else "generic_table",
                    "page_from": table.get("page_from"),
                    "page_to": table.get("page_to"),
                    "headers": table.get("headers", []),
                    "metadata_json": {
                        "document_id": document_id,
                        "title": raw.get("title"),
                        "category": metadata_row.get("category") or raw.get("category"),
                        "source_path": raw.get("source_path"),
                        "relative_path": relative_path,
                        "source_url": metadata_row.get("source_url") or raw.get("source_url"),
                    },
                    "status": status,
                },
            )
            table_count += 1

            for row_order, row in enumerate(table.get("rows", []), start=1):
                await knowledge_repo.create_knowledge_table_row(
                    db,
                    payload={
                        "table_pk_id": db_table.id,
                        "row_id": str(row.get("row_id", row_order)),
                        "label": row.get("label"),
                        "row_order": row_order,
                        "row_data": row,
                        "search_text": row.get("search_text")
                        or f"{table.get('title', '')} {row.get('label', '')} {row.get('amount_text', '')}".strip(),
                        "track_tags": row.get("track_tags", []),
                        "basis_tags": row.get("basis_tags", []),
                        "page_from": row.get("page_from", table.get("page_from")),
                        "page_to": row.get("page_to", table.get("page_to")),
                        "status": status,
                    },
                )
                row_count += 1

    return table_count, row_count


async def _sync_course_catalogs(
    db: AsyncSession,
    *,
    metadata_lookup: dict[str, dict[str, str]],
    source_version_lookup: dict[str, object],
) -> tuple[int, int]:
    catalog_count = 0
    course_count = 0

    for path in _course_catalog_source_files(metadata_lookup):
        relative_path = str(path.relative_to(PROJECT_ROOT)).replace("\\", "/")
        metadata_row = metadata_lookup.get(relative_path, {})
        status = metadata_row.get("status") or "draft"
        source_version_id = source_version_lookup.get(_source_id_from_relative_path(relative_path))
        if source_version_id is None:
            continue

        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet_schema_config = metadata_row.get("sheet_schema_config") or {}
        canonical_sheet_name = sheet_schema_config.get("canonical_sheet")
        sheet_roles = sheet_schema_config.get("sheet_roles") or {}
        canonical_sheet = None
        canonical_rows: list[list[object]] = []
        summary_sheet = None
        glossary_sheet = None

        for worksheet in workbook.worksheets:
            rows = _sheet_rows(worksheet)
            if not rows:
                continue
            headers = [_safe_cell(cell) for cell in rows[0]]
            if canonical_sheet_name and worksheet.title == canonical_sheet_name:
                canonical_sheet = worksheet.title
                canonical_rows = rows
            elif canonical_sheet is None and _looks_like_course_catalog_sheet(headers):
                canonical_sheet = worksheet.title
                canonical_rows = rows
            normalized_sheet_name = _normalize_search_text(worksheet.title)
            sheet_role = sheet_roles.get(worksheet.title)
            if summary_sheet is None and (sheet_role == "course_catalog_summary" or "tong quan" in normalized_sheet_name):
                summary_sheet = worksheet.title
            if glossary_sheet is None and (sheet_role == "course_catalog_glossary" or "ghi chu" in normalized_sheet_name):
                glossary_sheet = worksheet.title

        if canonical_sheet is None or not canonical_rows:
            continue

        header_map = {_normalize_search_text(header): index for index, header in enumerate(canonical_rows[0]) if header}
        course_name_index = header_map.get("ten hoc phan")
        course_code_index = header_map.get("ma hoc phan")
        credits_index = header_map.get("so tc")
        if course_name_index is None or course_code_index is None or credits_index is None:
            continue

        program_name = metadata_row.get("title") or path.stem.replace("_", " ").strip()
        catalog = await knowledge_repo.create_knowledge_course_catalog(
            db,
            payload={
                "catalog_id": f"catalog-{slugify(relative_path)}",
                "source_version_id": source_version_id,
                "program_name": program_name,
                "program_code": metadata_row.get("category"),
                "academic_year": metadata_row.get("version"),
                "canonical_sheet": canonical_sheet,
                "summary_sheet": summary_sheet,
                "glossary_sheet": glossary_sheet,
                "status": status,
                "metadata_json": {
                    "relative_path": relative_path,
                    "file_name": path.name,
                    "sheet_names": workbook.sheetnames,
                    "workbook_type": metadata_row.get("workbook_type"),
                    "sheet_schema_config": sheet_schema_config,
                },
            },
        )
        catalog_count += 1

        for row_offset, row in enumerate(canonical_rows[1:], start=2):
            course_name = _safe_cell(row[course_name_index]) if course_name_index < len(row) else ""
            course_code = _safe_cell(row[course_code_index]) if course_code_index < len(row) else ""
            credits = _parse_int(row[credits_index]) if credits_index < len(row) else None
            if not course_name or not course_code or credits is None:
                continue

            program_track_raw = _safe_cell(row[header_map.get("chuong trinh", -1)]) if "chuong trinh" in header_map and header_map["chuong trinh"] < len(row) else ""
            semester_label = _safe_cell(row[header_map.get("hoc ky", -1)]) if "hoc ky" in header_map and header_map["hoc ky"] < len(row) else "Khong ro hoc ky"
            program_track, degree_level = _parse_program_track(program_track_raw)
            normalized_course_name, aliases = _normalize_course_name(course_name)

            raw_row_data = {str(canonical_rows[0][index]): row[index] for index in range(min(len(canonical_rows[0]), len(row))) if canonical_rows[0][index]}
            search_parts = [
                program_track_raw,
                semester_label,
                course_name,
                normalized_course_name,
                course_code,
                str(credits),
                _safe_cell(row[header_map.get("nngd", -1)]) if "nngd" in header_map and header_map["nngd"] < len(row) else "",
                _safe_cell(row[header_map.get("hp tien quyet", -1)]) if "hp tien quyet" in header_map and header_map["hp tien quyet"] < len(row) else "",
            ]

            await knowledge_repo.create_knowledge_course(
                db,
                payload={
                    "course_record_id": f"{catalog.catalog_id}-row-{row_offset:04d}",
                    "catalog_pk_id": catalog.id,
                    "sheet_name": canonical_sheet,
                    "row_number": row_offset,
                    "program_track": program_track,
                    "degree_level": degree_level,
                    "program_variant": None,
                    "semester_number": _parse_semester_number(semester_label),
                    "semester_label": semester_label,
                    "course_order": _parse_int(row[header_map.get("tt", -1)]) if "tt" in header_map and header_map["tt"] < len(row) else None,
                    "course_name": course_name,
                    "normalized_course_name": normalized_course_name,
                    "course_name_aliases": aliases,
                    "course_code": course_code,
                    "credits": credits,
                    "lecture_hours": _parse_int(row[header_map.get("lt", -1)]) if "lt" in header_map and header_map["lt"] < len(row) else None,
                    "discussion_hours": _parse_int(row[header_map.get("tl bt", -1)]) if "tl bt" in header_map and header_map["tl bt"] < len(row) else None,
                    "course_design_hours": _parse_int(row[header_map.get("tkmh", -1)]) if "tkmh" in header_map and header_map["tkmh"] < len(row) else None,
                    "project_hours": _parse_int(row[header_map.get("btl", -1)]) if "btl" in header_map and header_map["btl"] < len(row) else None,
                    "lab_hours": _parse_int(row[header_map.get("tn", -1)]) if "tn" in header_map and header_map["tn"] < len(row) else None,
                    "practice_hours": _parse_int(row[header_map.get("th", -1)]) if "th" in header_map and header_map["th"] < len(row) else None,
                    "self_study_hours": _parse_int(row[header_map.get("thoc", -1)]) if "thoc" in header_map and header_map["thoc"] < len(row) else None,
                    "prerequisite_text": _safe_cell(row[header_map.get("hp tien quyet", -1)]) if "hp tien quyet" in header_map and header_map["hp tien quyet"] < len(row) else None,
                    "prerequisite_codes": [],
                    "teaching_language": _parse_teaching_language(_safe_cell(row[header_map.get("nngd", -1)])) if "nngd" in header_map and header_map["nngd"] < len(row) else None,
                    "search_text": " ".join(part for part in search_parts if part).strip(),
                    "raw_row_data": raw_row_data,
                    "status": status,
                },
            )
            course_count += 1

    return catalog_count, course_count


async def _sync_faqs(db: AsyncSession) -> int:
    count = 0
    if not FAQS_DIR.exists():
        return count

    for faq in list_faqs():
        await knowledge_repo.create_knowledge_faq(
            db,
            payload={
                "faq_id": faq.id,
                "source_version_id": None,
                "question": faq.question,
                "answer": faq.answer,
                "category": faq.category,
                "subcategory": None,
                "keywords": faq.keywords,
                "trust_level": "reviewed_faq",
                "status": "published",
                "source_url": faq.source_url,
                "last_reviewed_at": faq.updated_at,
                "reviewed_by": None,
            },
        )
        count += 1

    return count


async def _sync_forms(db: AsyncSession) -> int:
    count = 0
    if not FORMS_DIR.exists():
        return count

    for form in list_forms():
        field_payloads = []
        for field in form.fields:
            if hasattr(field, "model_dump"):
                field_payloads.append(field.model_dump())
            elif hasattr(field, "dict"):
                field_payloads.append(field.dict())
            else:
                field_payloads.append(field)
        await knowledge_repo.create_knowledge_form(
            db,
            payload={
                "form_id": form.id,
                "title": form.title,
                "category": form.category,
                "description": form.description,
                "keywords": form.keywords,
                "fields": field_payloads,
                "print_template_type": "administrative_letter",
                "related_procedure_id": None,
                "source_ids": [],
                "status": "published",
            },
        )
        count += 1

    return count


async def _sync_procedures(db: AsyncSession) -> int:
    count = 0
    for procedure in PROCEDURE_CATALOG:
        matched_form = _find_admin_form(procedure.get("form_topic", ""), procedure["title"])
        category = None
        if isinstance(matched_form, dict):
            category = matched_form.get("category")
        if not category:
            category = "Thủ tục hành chính"
        await knowledge_repo.create_knowledge_procedure(
            db,
            payload={
                "procedure_id": procedure["id"],
                "title": procedure["title"],
                "category": category,
                "keywords": procedure.get("keywords", []),
                "eligibility": procedure.get("eligibility", []),
                "required_documents": procedure.get("required_documents", []),
                "steps": procedure.get("steps", []),
                "contact_office": procedure.get("contact_office"),
                "related_form_id": matched_form.get("id") if isinstance(matched_form, dict) else None,
                "source_ids": [],
                "status": "published",
                "trust_level": "official",
            },
        )
        count += 1

    return count


async def sync_normalized_knowledge(db: AsyncSession) -> dict[str, int]:
    """Sync normalized-layer tables from processed artifacts and admin content."""
    if not await _normalized_tables_ready(db):
        return {
            "chunks": 0,
            "tables": 0,
            "table_rows": 0,
            "faqs": 0,
            "forms": 0,
            "procedures": 0,
            "course_catalogs": 0,
            "courses": 0,
        }

    manifest_lookup = _manifest_by_document_id()
    metadata_lookup = load_document_metadata_map()
    source_version_lookup = await _active_source_versions_by_source_id(db)

    await knowledge_repo.clear_normalized_knowledge(db)

    chunk_count = await _sync_chunks(
        db,
        manifest_lookup=manifest_lookup,
        metadata_lookup=metadata_lookup,
        source_version_lookup=source_version_lookup,
    )
    table_count, row_count = await _sync_tables(
        db,
        manifest_lookup=manifest_lookup,
        metadata_lookup=metadata_lookup,
        source_version_lookup=source_version_lookup,
    )
    course_catalog_count, course_count = await _sync_course_catalogs(
        db,
        metadata_lookup=metadata_lookup,
        source_version_lookup=source_version_lookup,
    )
    faq_count = await _sync_faqs(db)
    form_count = await _sync_forms(db)
    procedure_count = await _sync_procedures(db)

    return {
        "chunks": chunk_count,
        "tables": table_count,
        "table_rows": row_count,
        "faqs": faq_count,
        "forms": form_count,
        "procedures": procedure_count,
        "course_catalogs": course_catalog_count,
        "courses": course_count,
    }
